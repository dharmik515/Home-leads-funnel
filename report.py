"""
Builds the downloadable Excel **summary report** — a colour-coded, automated
replacement for the hand-built Summary sheet.

Sheets:
  - Summary: dated header, colourful KPI cards, and a channel-wise distribution
    table with **count + % side by side** for every funnel stage. Each status owns
    a colour (green/red/amber/blue) carried through headers, column tints, KPI cards
    and charts. Data bars on Total Leads, a red→green colour scale on Conversion %,
    plus a stacked bar chart (funnel per channel) and an outcome pie.
  - By Period: the chosen time bucket (weekly … annual) with counts, % and a chart.
  - Rejection Reasons: ranked reasons with share + bar chart.

The layout is driven by data.STATUS_ORDER, so adding/removing a funnel bucket
(e.g. Awaiting Confirmation) flows through automatically.
"""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import data as D

# --- palette -------------------------------------------------------------
SLATE = "1F2937"
INK = "0F172A"
MUTE = "64748B"

# Status colour system: strong (headers/cards) + tint (column bands).
STATUS_HEX = {"Done": "16A34A", "Rejected": "DC2626",
              "Pending": "F59E0B", "Awaiting Confirmation": "2563EB"}
STATUS_TINT = {"Done": "DCFCE7", "Rejected": "FEE2E2",
               "Pending": "FEF3C7", "Awaiting Confirmation": "DBEAFE"}
DISPLAY = {"Done": "Deals Done"}   # status -> column label override

# --- reusable styles -----------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor=SLATE)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(size=18, bold=True, color=INK)
DATE_FONT = Font(size=10, bold=True, color="FFFFFF")
SUB_FONT = Font(size=10, italic=True, color=MUTE)
TOTAL_FILL = PatternFill("solid", fgColor="E2E8F0")
TOTAL_FONT = Font(bold=True, color=INK)
BAND_FILL = PatternFill("solid", fgColor="EFF3F8")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
CANVAS_FILL = PatternFill("solid", fgColor="E7ECF3")   # light dashboard background
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
_thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CNT_FMT = "#,##0"
PCT_FMT = "0.0%"


def _ref(ws, c1, r1, c2, r2) -> str:
    """A quoted A1-style range string for the given sheet/cells."""
    a = f"${get_column_letter(c1)}${r1}"
    b = f"${get_column_letter(c2)}${r2}"
    return f"'{ws.title}'!{a}:{b}"


def _text_categories(chart, ref: str):
    """Set the chart's category axis from a *text* range and make axes visible.

    Two openpyxl gotchas fixed here:
    1. set_categories() writes a numeric reference, which Excel can't read for
       text labels (they render blank) -> force a strRef.
    2. Chart axes default to delete=True, so Excel hides the axis labels
       entirely -> force both axes visible.
    """
    src = AxDataSource(strRef=StrRef(f=ref))
    for s in chart.series:
        s.cat = src
    chart.x_axis.delete = False
    chart.y_axis.delete = False


def _value_labels(num_fmt: str | None = None) -> DataLabelList:
    """Data labels that show ONLY the value (not the noisy series/category name)."""
    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.showBubbleSize = False
    if num_fmt:
        dl.numFmt = num_fmt
    return dl


def build_report_xlsx(df, bucket: str = "Monthly") -> bytes:
    funnel = D.funnel_counts(df)
    summ = D.channel_summary(df)
    ppartner = D.period_partner(df, bucket)
    rpivot = D.reject_pivot(df, top=8)
    dmin, dmax = df["_date"].min(), df["_date"].max()
    period = (f"{dmin:%d %b %Y} – {dmax:%d %b %Y}"
              if pd.notna(dmin) and pd.notna(dmax) else "all dates")

    wb = Workbook()
    _summary_sheet(wb.active, df, funnel, summ, period)
    _period_sheet(wb.create_sheet("By Period"), ppartner, bucket, period)
    _reasons_sheet(wb.create_sheet("Rejection Reasons"), rpivot, funnel, period)
    remarks = D.appraisal_remarks(df)
    if not remarks.empty:
        _remarks_sheet(wb.create_sheet("Appraiser Remarks"), remarks, period)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _remarks_sheet(ws, remarks, period):
    """Free-text appraiser remarks from Fero, one row per device."""
    _dated_header(ws, "Appraiser Remarks (from Fero)", period)
    headers = ["Deal ID", "Channel", "Status", "Reason", "Remark"]
    widths = [22, 22, 12, 34, 50]
    top = 4
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=top, column=j, value=h)
        cell.font = HEADER_FONT; cell.alignment = LEFT if j in (1, 4, 5) else CENTER
        cell.border = BORDER; cell.fill = HEADER_FILL
    r = top
    for _, row in remarks.iterrows():
        r += 1
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=r, column=j, value=str(row[h]))
            cell.border = BORDER
            cell.alignment = LEFT if j in (1, 4, 5) else CENTER
            if (r - top) % 2 == 0:
                cell.fill = BAND_FILL
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = f"A{top+1}"


# -------------------------------------------------------------------------
def _dated_header(ws, title, period):
    """Title + a coloured dated ribbon (the 'date in the output')."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"  Data period: {period}      Report generated: {date.today():%d %b %Y}"
    ws["A2"].font = DATE_FONT
    ws["A2"].fill = PatternFill("solid", fgColor=SLATE)
    ws["A2"].alignment = LEFT
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 20


def _summary_sheet(ws, df, funnel, summ, period):
    ws.title = "Summary"
    # Light dashboard canvas behind everything; cards/table/charts sit on top.
    for rr in range(1, 64):
        for cc in range(1, 16):
            ws.cell(row=rr, column=cc).fill = CANVAS_FILL
    _dated_header(ws, "Home Leads Funnel — Summary", period)
    ws["G1"] = f"{funnel['Total Leads']:,} leads · {df['Channel'].nunique()} channels"
    ws["G1"].font = SUB_FONT

    statuses = D.STATUS_ORDER
    total = funnel["Total Leads"] or 1

    # ---- colourful KPI cards (row 4 label / row 5 value), 2 cols each ----
    cards = [("Total Leads", funnel["Total Leads"], None, "4F46E5")]
    for s in statuses:
        key = "Deals Done" if s == "Done" else s
        cards.append((key, funnel.get("Deals Done" if s == "Done" else s, 0),
                      funnel.get("Deals Done" if s == "Done" else s, 0) / total, STATUS_HEX[s]))
    for i, (lbl, val, pct, hexcol) in enumerate(cards):
        c0 = 1 + i * 2
        ws.merge_cells(start_row=4, start_column=c0, end_row=4, end_column=c0 + 1)
        ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 1)
        lc = ws.cell(row=4, column=c0, value=lbl.upper())
        lc.fill = PatternFill("solid", fgColor=hexcol)
        lc.font = Font(bold=True, color="FFFFFF", size=9)
        lc.alignment = CENTER
        txt = f"{val:,}" + (f"   {pct:.0%}" if pct is not None else "")
        vc = ws.cell(row=5, column=c0, value=txt)
        vc.fill = PatternFill("solid", fgColor=hexcol)
        vc.font = Font(bold=True, color="FFFFFF", size=15)
        vc.alignment = CENTER
        ws.row_dimensions[5].height = 26

    # ---- channel distribution table (count + % side by side) ----
    # Dynamic columns: Channel | Total | [status count | %]... | Conversion %
    spec = [("Channel", "text", None), ("Total Leads", "count", None)]
    for s in statuses:
        spec.append((DISPLAY.get(s, s), "count", s))
        spec.append(("%", "pct", s))
    spec.append(("Conversion %", "pct", None))
    ncol = len(spec)
    total_col = 2
    conv_col = ncol

    top = 8
    for j, (label, _kind, s) in enumerate(spec, 1):
        cell = ws.cell(row=top, column=j, value=label)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        cell.fill = PatternFill("solid", fgColor=STATUS_HEX[s]) if s else HEADER_FILL

    def status_count(row_or_funnel, s, total_leads):
        col = "Done" if s == "Done" else s
        return row_or_funnel[col]

    body = summ[summ["Channel"] != "ALL CHANNELS"]
    r = top
    for _, row in body.iterrows():
        r += 1
        tot = row["Total Leads"] or 1
        for j, (label, kind, s) in enumerate(spec, 1):
            if kind == "text":
                v = row["Channel"]
            elif label == "Total Leads":
                v = int(row["Total Leads"])
            elif label == "Conversion %":
                v = row["Done"] / tot
            elif kind == "count":
                v = int(status_count(row, s, tot))
            else:  # pct beside its status count
                v = status_count(row, s, tot) / tot
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = LEFT if kind == "text" else CENTER
            if kind == "pct":
                cell.number_format = PCT_FMT
            elif kind == "count" and label != "Channel":
                cell.number_format = CNT_FMT
            # colour: status columns get their tint; others get crisp banding
            if s:
                cell.fill = PatternFill("solid", fgColor=STATUS_TINT[s])
            else:
                cell.fill = BAND_FILL if (r - top) % 2 == 0 else WHITE_FILL

    # total row
    r += 1
    tot_all = funnel["Total Leads"] or 1
    for j, (label, kind, s) in enumerate(spec, 1):
        if kind == "text":
            v = "ALL CHANNELS"
        elif label == "Total Leads":
            v = funnel["Total Leads"]
        elif label == "Conversion %":
            v = funnel["Deals Done"] / tot_all
        elif kind == "count":
            v = funnel["Deals Done" if s == "Done" else s]
        else:
            v = funnel["Deals Done" if s == "Done" else s] / tot_all
        cell = ws.cell(row=r, column=j, value=v)
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.border = BORDER
        cell.alignment = LEFT if kind == "text" else CENTER
        if kind == "pct":
            cell.number_format = PCT_FMT
        elif kind == "count" and label != "Channel":
            cell.number_format = CNT_FMT

    last_data_row = r - 1
    L = get_column_letter
    # data bars on Total Leads + colour scale on Conversion %  (engagement)
    ws.conditional_formatting.add(
        f"{L(total_col)}{top+1}:{L(total_col)}{last_data_row}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="6366F1"))
    ws.conditional_formatting.add(
        f"{L(conv_col)}{top+1}:{L(conv_col)}{last_data_row}",
        ColorScaleRule(start_type="num", start_value=0, start_color="F8696B",
                       mid_type="percentile", mid_value=50, mid_color="FFEB84",
                       end_type="max", end_color="63BE7B"))

    # widths + freeze
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    for j in range(3, ncol):
        ws.column_dimensions[L(j)].width = 11 if (j % 2 == 1) else 8
    ws.column_dimensions[L(conv_col)].width = 13
    ws.freeze_panes = f"A{top+1}"

    _charts(ws, statuses, top, last_data_row, conv_col, anchor_row=r + 3)


def _charts(ws, statuses, top, last_data_row, conv_col, anchor_row):
    """Two charts, sourced from the VISIBLE table (Excel won't plot hidden cells).
    Channel labels via strRef; funnel uses 100%-stacked so small channels aren't
    crushed by the highest-volume one."""
    cat_ref = _ref(ws, 1, top + 1, 1, last_data_row)          # channel names (text)

    # 1) Funnel mix by channel (100% stacked → comparable regardless of volume).
    bar = BarChart()
    bar.type = "bar"; bar.grouping = "percentStacked"; bar.overlap = 100
    bar.title = "Funnel mix by channel (share of leads)"
    bar.height = 9.5; bar.width = 20
    for k, s in enumerate(statuses):
        ref = Reference(ws, min_col=3 + 2 * k, min_row=top, max_row=last_data_row)
        bar.add_data(ref, titles_from_data=True)
    for srs, s in zip(bar.series, statuses):
        srs.graphicalProperties.solidFill = STATUS_HEX[s]
        srs.graphicalProperties.line.solidFill = STATUS_HEX[s]
    _text_categories(bar, cat_ref)
    bar.y_axis.numFmt = "0%"               # value axis = share of leads
    bar.legend.position = "b"
    ws.add_chart(bar, f"A{anchor_row}")

    # 2) Conversion % by channel — horizontal so channel names sit on the left.
    conv = BarChart()
    conv.type = "bar"; conv.title = "Conversion % by channel"
    conv.height = 9.5; conv.width = 20
    conv.add_data(Reference(ws, min_col=conv_col, min_row=top, max_row=last_data_row),
                  titles_from_data=True)
    conv.series[0].graphicalProperties.solidFill = "4F46E5"
    _text_categories(conv, cat_ref)
    conv.y_axis.numFmt = "0%"
    conv.dataLabels = _value_labels("0%")
    conv.legend = None
    ws.add_chart(conv, f"A{anchor_row + 21}")


# -------------------------------------------------------------------------
# 12 distinct colours + cycling markers so every partner line stays identifiable.
LINE_PALETTE = ["2563EB", "16A34A", "DC2626", "D97706", "7C3AED", "0891B2",
                "DB2777", "65A30D", "EA580C", "0D9488", "9333EA", "475569"]
MARKERS = ["circle", "square", "triangle", "diamond", "x", "plus"]


def _period_sheet(ws, pp, bucket, period):
    """Partner-wise: leads per time bucket × partner (heatmap matrix + trend lines)."""
    _dated_header(ws, f"Leads by {bucket} × Partner", period)
    if pp is None or pp.empty:
        ws["A4"] = "No dated deals in this selection."
        ws["A4"].font = SUB_FONT
        return

    partners = list(pp.columns)               # already ordered busiest-first
    periods = list(pp.index)
    ncol = 1 + len(partners) + 1              # Period | partners... | Total
    total_col = ncol
    top = 4

    # ---- header (Period | rotated partner names | Total) ----
    hp = ws.cell(row=top, column=1, value="Period")
    hp.font = HEADER_FONT; hp.alignment = LEFT; hp.border = BORDER; hp.fill = HEADER_FILL
    for k, ch in enumerate(partners, start=2):
        cell = ws.cell(row=top, column=k, value=ch)
        cell.font = HEADER_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
        cell.fill = PatternFill("solid", fgColor="4F46E5")
    tc = ws.cell(row=top, column=total_col, value="Total")
    tc.font = HEADER_FONT; tc.alignment = CENTER; tc.border = BORDER
    tc.fill = PatternFill("solid", fgColor="312E81")
    ws.row_dimensions[top].height = 96

    # ---- body: one row per period ----
    r = top
    body_first = top + 1
    for prd in periods:
        r += 1
        a = ws.cell(row=r, column=1, value=prd); a.border = BORDER; a.alignment = LEFT
        a.fill = BAND_FILL if (r - top) % 2 == 0 else WHITE_FILL
        row_total = 0
        for k, ch in enumerate(partners, start=2):
            v = int(pp.loc[prd, ch]); row_total += v
            cell = ws.cell(row=r, column=k, value=v)
            cell.border = BORDER; cell.alignment = CENTER; cell.number_format = CNT_FMT
        tcell = ws.cell(row=r, column=total_col, value=row_total)
        tcell.border = BORDER; tcell.alignment = CENTER; tcell.number_format = CNT_FMT
    body_last = r

    # ---- total row (per partner) ----
    r += 1
    grand = ws.cell(row=r, column=1, value="TOTAL")
    grand.fill = TOTAL_FILL; grand.font = TOTAL_FONT; grand.border = BORDER; grand.alignment = LEFT
    for k, ch in enumerate(partners, start=2):
        v = int(pp[ch].sum())
        cell = ws.cell(row=r, column=k, value=v)
        cell.fill = TOTAL_FILL; cell.font = TOTAL_FONT; cell.border = BORDER
        cell.alignment = CENTER; cell.number_format = CNT_FMT
    gt = ws.cell(row=r, column=total_col, value=int(pp.values.sum()))
    gt.fill = TOTAL_FILL; gt.font = TOTAL_FONT; gt.border = BORDER; gt.alignment = CENTER; gt.number_format = CNT_FMT

    # ---- heatmap over matrix + data bars on Total column ----
    L = get_column_letter
    ws.conditional_formatting.add(
        f"{L(2)}{body_first}:{L(total_col-1)}{body_last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       mid_type="percentile", mid_value=60, mid_color="C7D2FE",
                       end_type="max", end_color="4F46E5"))
    ws.conditional_formatting.add(
        f"{L(total_col)}{body_first}:{L(total_col)}{body_last}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="312E81"))

    ws.column_dimensions["A"].width = 16
    for k in range(2, total_col):
        ws.column_dimensions[L(k)].width = 6
    ws.column_dimensions[L(total_col)].width = 9
    ws.freeze_panes = "B5"

    # ---- trend chart: every partner as its own line (distinct colour + marker) ----
    line = LineChart()
    line.title = "Lead trend by partner"
    line.height = 11; line.width = 26
    line.y_axis.title = "Leads"
    line.x_axis.delete = False; line.y_axis.delete = False
    for i in range(len(partners)):
        ref = Reference(ws, min_col=2 + i, min_row=top, max_row=body_last)
        line.add_data(ref, titles_from_data=True)
    _text_categories(line, _ref(ws, 1, body_first, 1, body_last))   # period labels (text)
    for i, srs in enumerate(line.series):
        colour = LINE_PALETTE[i % len(LINE_PALETTE)]
        srs.graphicalProperties.line.solidFill = colour
        srs.graphicalProperties.line.width = 26000   # ~2pt
        srs.marker = Marker(symbol=MARKERS[i % len(MARKERS)], size=6)
        srs.marker.graphicalProperties.solidFill = colour
        srs.smooth = False
    line.legend.position = "r"          # right side: doesn't collide with date labels
    ws.add_chart(line, f"A{r + 3}")


def _reasons_sheet(ws, pivot, funnel, period):
    _dated_header(ws, "Rejection Reasons by Partner", period)
    if pivot is None or pivot.empty:
        ws["A4"] = "No rejections with a recorded reason in this selection."
        ws["A4"].font = SUB_FONT
        return
    top = 4
    channels = [c for c in pivot.columns if c != "Total"]
    reasons = list(pivot.index)                 # includes 'ALL REASONS' total row last
    ncol = 1 + len(channels) + 1                # Reason | channels... | Total
    total_col = ncol

    # ---- header (reason | rotated channel names | total) ----
    h = ws.cell(row=top, column=1, value="Rejection reason")
    h.font = HEADER_FONT; h.alignment = LEFT; h.border = BORDER
    h.fill = PatternFill("solid", fgColor="DC2626")
    for k, ch in enumerate(channels, start=2):
        cell = ws.cell(row=top, column=k, value=ch)
        cell.font = HEADER_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="bottom", text_rotation=90)
        cell.fill = HEADER_FILL
    tc = ws.cell(row=top, column=total_col, value="Total")
    tc.font = HEADER_FONT; tc.alignment = CENTER; tc.border = BORDER; tc.fill = PatternFill("solid", fgColor="991B1B")
    ws.row_dimensions[top].height = 96

    # ---- body ----
    r = top
    body_first = top + 1
    for reason in reasons:
        r += 1
        is_total = reason == "ALL REASONS"
        a = ws.cell(row=r, column=1, value=reason)
        a.border = BORDER; a.alignment = LEFT
        for k, ch in enumerate(channels, start=2):
            cell = ws.cell(row=r, column=k, value=int(pivot.loc[reason, ch]))
            cell.border = BORDER; cell.alignment = CENTER; cell.number_format = CNT_FMT
        t = ws.cell(row=r, column=total_col, value=int(pivot.loc[reason, "Total"]))
        t.border = BORDER; t.alignment = CENTER; t.number_format = CNT_FMT
        if is_total:
            for cc in range(1, ncol + 1):
                ws.cell(row=r, column=cc).fill = TOTAL_FILL
                ws.cell(row=r, column=cc).font = TOTAL_FONT
        else:
            a.fill = BAND_FILL if (r - top) % 2 == 0 else WHITE_FILL
    body_last = r - 1            # excludes ALL REASONS row
    grand_total_row = r

    # ---- heatmap over the matrix body (channel cells) + data bars on Total ----
    L = get_column_letter
    ws.conditional_formatting.add(
        f"{L(2)}{body_first}:{L(total_col-1)}{body_last}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                       mid_type="percentile", mid_value=60, mid_color="FCA5A5",
                       end_type="max", end_color="DC2626"))
    ws.conditional_formatting.add(
        f"{L(total_col)}{body_first}:{L(total_col)}{body_last}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color="991B1B"))

    # ---- widths + freeze (keep reason column + header visible) ----
    ws.column_dimensions["A"].width = 30
    for k in range(2, total_col):
        ws.column_dimensions[L(k)].width = 5.5
    ws.column_dimensions[L(total_col)].width = 8
    ws.freeze_panes = "B5"

    # ---- two simple, single-message charts (easier to read than a stack) ----
    # A) Top rejection reasons overall (horizontal bar) — uses the Total column.
    rbar = BarChart()
    rbar.type = "bar"; rbar.title = "Top rejection reasons (all partners) — Source: Fero"
    rbar.height = 0.55 * (body_last - body_first + 1) + 2.5; rbar.width = 18
    rbar.add_data(Reference(ws, min_col=total_col, min_row=body_first, max_row=body_last))
    _text_categories(rbar, _ref(ws, 1, body_first, 1, body_last))     # reason names (text)
    rbar.series[0].graphicalProperties.solidFill = "DC2626"
    rbar.dataLabels = _value_labels()
    rbar.legend = None
    ws.add_chart(rbar, f"A{grand_total_row + 3}")

    # B) Total rejections by partner (horizontal — partner names on the left).
    pbar = BarChart()
    pbar.type = "bar"; pbar.title = "Total rejections by partner — Source: Home Funnel"
    pbar.height = 9; pbar.width = 18
    data = Reference(ws, min_col=2, max_col=1 + len(channels), min_row=grand_total_row, max_row=grand_total_row)
    pbar.add_data(data, from_rows=True)
    _text_categories(pbar, _ref(ws, 2, top, 1 + len(channels), top))   # partner names (text, header row)
    pbar.series[0].graphicalProperties.solidFill = "991B1B"
    pbar.dataLabels = _value_labels()
    pbar.legend = None
    ws.add_chart(pbar, f"A{grand_total_row + 22}")
